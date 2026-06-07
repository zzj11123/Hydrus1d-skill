#!/usr/bin/env python3
"""Validate HYDRUS-1D observation tables.

Reads CSV by default. XLSX is supported when openpyxl is installed.
Outputs JSON so Codex can consume the result reliably.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from pathlib import Path
from typing import Any


ALIASES = {
    "time": {"time", "t", "day", "days", "时间", "天数"},
    "depth_cm": {"depth_cm", "depth", "z", "x", "cm", "深度", "土层深度", "观测深度"},
    "variable": {"variable", "var", "type", "指标", "变量", "项目"},
    "observed": {"observed", "obs", "value", "measured", "实测值", "观测值", "测定值"},
    "unit": {"unit", "units", "单位"},
    "solute_id": {"solute_id", "solute", "solute_index", "溶质编号", "盐分编号"},
    "species": {"species", "ion", "salt", "solute_name", "离子", "盐分", "溶质名称"},
    "replicate": {"replicate", "rep", "重复", "重复号"},
    "weight": {"weight", "权重"},
    "group": {"group", "组", "分组"},
    "note": {"note", "notes", "备注", "说明"},
}

SUPPORTED_VARIABLES = {
    "theta",
    "h",
    "water_content",
    "pressure_head",
    "top_flux",
    "bottom_flux",
    "drainage",
    "cumulative_infiltration",
    "concentration",
    "solute1",
    "solute2",
    "solute3",
    "ec",
    "tds",
    "salt",
}

REQUIRED = ["time", "depth_cm", "variable", "observed", "unit"]


def norm(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t") if sample.strip() else csv.excel
        return list(csv.DictReader(handle, dialect=dialect))


def read_xlsx(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover - environment dependent
        raise SystemExit(f"XLSX input requires openpyxl: {exc}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if value is None else str(value) for value in rows[0]]
    records = []
    for row in rows[1:]:
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return records


def map_columns(headers: list[str]) -> tuple[dict[str, str], dict[str, list[str]]]:
    mapping: dict[str, str] = {}
    ambiguous: dict[str, list[str]] = {}
    normalized = {header: norm(header) for header in headers}
    for canonical, aliases in ALIASES.items():
        matches = [header for header, value in normalized.items() if value in {norm(alias) for alias in aliases}]
        if len(matches) == 1:
            mapping[canonical] = matches[0]
        elif len(matches) > 1:
            ambiguous[canonical] = matches
    return mapping, ambiguous


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def validate(path: Path, sheet: str | None = None) -> dict[str, Any]:
    records = read_xlsx(path, sheet) if path.suffix.lower() in {".xlsx", ".xlsm"} else read_csv(path)
    headers = list(records[0].keys()) if records else []
    mapping, ambiguous = map_columns(headers)
    missing = [field for field in REQUIRED if field not in mapping]
    issues: list[str] = []
    warnings: list[str] = []
    if missing:
        issues.append(f"Missing required fields: {', '.join(missing)}")
    if ambiguous:
        warnings.append(f"Ambiguous fields: {ambiguous}")

    variable_counts: dict[str, int] = {}
    bad_rows = 0
    solute_like_without_id = 0
    time_values: list[float] = []
    depth_values: list[float] = []

    for index, row in enumerate(records, start=2):
        if missing:
            continue
        time_value = as_float(row.get(mapping["time"]))
        depth_value = as_float(row.get(mapping["depth_cm"]))
        observed_value = as_float(row.get(mapping["observed"]))
        variable = str(row.get(mapping["variable"], "")).strip()
        variable_key = norm(variable)

        if time_value is None or depth_value is None or observed_value is None or not variable:
            bad_rows += 1
            continue
        time_values.append(time_value)
        depth_values.append(depth_value)
        variable_counts[variable_key] = variable_counts.get(variable_key, 0) + 1
        if variable_key not in SUPPORTED_VARIABLES:
            warnings.append(f"Row {index}: variable '{variable}' is not a standard HYDRUS mapping.")
        if variable_key in {"concentration", "ec", "tds", "salt"} or variable_key.startswith("solute"):
            if "solute_id" not in mapping and "species" not in mapping:
                solute_like_without_id += 1

    if bad_rows:
        issues.append(f"Rows with missing/non-numeric required values: {bad_rows}")
    if solute_like_without_id:
        warnings.append("Solute/salinity observations should include solute_id or species.")

    return {
        "path": str(path),
        "rows": len(records),
        "headers": headers,
        "mapping": mapping,
        "ambiguous": ambiguous,
        "missing_required": missing,
        "issues": issues,
        "warnings": sorted(set(warnings)),
        "variables": variable_counts,
        "time_min": min(time_values) if time_values else None,
        "time_max": max(time_values) if time_values else None,
        "depth_min_cm": min(depth_values) if depth_values else None,
        "depth_max_cm": max(depth_values) if depth_values else None,
        "valid": not issues and not missing,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate HYDRUS-1D observation CSV/XLSX files.")
    parser.add_argument("path", help="CSV or XLSX observation file.")
    parser.add_argument("--sheet", help="Worksheet name for XLSX files.")
    args = parser.parse_args()
    result = validate(Path(args.path), args.sheet)
    json.dump(result, sys.stdout, ensure_ascii=False, indent=2)
    print()
    return 0 if result["valid"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
